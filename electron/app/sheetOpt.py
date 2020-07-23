import os
import sys
import openpyxl
from openpyxl import Workbook
from datetime import datetime

E9PASS = 'Desktop\\E9pass'

def checkBackUpDir():
    home = os.environ['USERPROFILE']
    dirPath = os.path.join(home, E9PASS)
    if os.path.isdir(dirPath):
        return dirPath
    else:
        os.makedirs(dirPath)
        return dirPath

def isFileExists():
    checkBackUpDir()
    home = os.environ['USERPROFILE']
    backUpDir = os.path.join(home, E9PASS)
    os.chdir(backUpDir)
    if os.path.isfile('Daily Uploads.xlsx'):
        return True
    else:
        return False

def readFromFile():
    result = isFileExists()
    if result:
        home = os.environ['USERPROFILE']
        backUpDir = os.path.join(home, E9PASS)
        os.chdir(backUpDir)
        wb = openpyxl.load_workbook('Daily Uploads.xlsx')
        sheet = wb.active
        rowCount = sheet.max_row
        sheetData = []
        for i in range(2, rowCount + 1):
            try:
                rowData = {
                    "appNumber" : str(sheet.cell(row=i, column=1).value),
                    "arcNumber" : str(sheet.cell(row=i, column=2).value),
                    "userName" : str(sheet.cell(row=i, column=3).value),
                    "passStatus": str(sheet.cell(row=i, column=4).value),
                    "date" : str(sheet.cell(row=i, column=5).value),
                }
                sheetData.append(rowData)
            except:
                pass
        return sheetData
    else:
        return None

def writeToFile(sheetData):
    home = os.environ['USERPROFILE']
    backUpDir = os.path.join(home, E9PASS)
    os.chdir(backUpDir)
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'sheet1'
    titles = ['Application NO', 'ARC Number', 'Registered Name', 'Pass Status', 'Completed Date', 'Certificate', 'Pdf', 'Zip']
    sheet.append(titles)
    for row in sheetData:
        rowData = (row['appNumber'], row['arcNumber'], row['userName'], row['passStatus'], row['date'], row['certStatus'], row['pdfStatus'], row['zipStatus'])
        sheet.append(rowData)
    now = datetime.now()
    date_time = now.strftime("%m%d%Y%H%M%S")
    fileName = date_time + '_' + str(len(sheetData)) + '.xlsx'
    wb.save(fileName)
    return True

def appendToFile(appendData):
    oldData = readFromFile()
    if bool(oldData):
        result = writeToFile(oldData + appendData)
        return result
    else:
        result = writeToFile(appendData)
        return result
